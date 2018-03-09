# original C:\AgilyticsCharu\PyExe-Charu\Trials\writeRecordsInExcel.py
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color


# list_Consultant=[{'id': 3, 'RFPName': ('1A', '1B', '1C', '2A', '2B', 3, None, None), 'RFPMarks': (21, 3, 1, 22, 24, 20, -1, -1), 'TotalMarks': 89},
#                  {'id': 2, 'RFPName': ('1A', '1B', '1C', 3, None, None), 'RFPMarks': (21, 3, 1, 20, -1, -1), 'TotalMarks': 43},
#                  {'id': 5, 'RFPName': ('1B', '1C', '1D', 3, None, None), 'RFPMarks': (3, 1, 7, 20, -1, -1), 'TotalMarks': 29},
#                  {'id': 1, 'RFPName': ('1B', '1C', 3, None, None), 'RFPMarks': (3, 1, 25, -1, -1), 'TotalMarks': 27},
#                  {'id': 6, 'RFPName': ('1B', '1C'), 'RFPMarks': (3, 1), 'TotalMarks': 4},
#                  {'id': 4, 'RFPName': ('1B', '1C', None, None), 'RFPMarks': (3, 1, -1, -1), 'TotalMarks': 2}]
# list_Consultant=[{'id': 3, 'RFPName': ('1A', '1B', '1C', '2A', '2B', 3), 'RFPMarks': (21, 3, 1, 22, 24, 20), 'TotalMarks': 91, 'name': 'SUBRATA MUKHOPADHYAY',  'email': 'm.subrata59@gmail.com', 'state': 'DELHI', 'district': 'District New Delhi', 'address': '-', 'pan': 'AHAPM6089L', 'mobile': '9434423129', 'alter_mobile': '0', 'landline': '9434423129'},
#  {'id': 2, 'RFPName': ('1A', '1B', '1C', 3), 'RFPMarks': (21, 3, 1, 20), 'TotalMarks': 45, 'name': 'Rakesh Arora', 'email': 'aroraselfie@gmail.com', 'state': 'PUNJAB', 'district': 'Muktsar', 'address': '332, PUDA Colony, Malout, Punjab', 'pan': 'AAMPA2916M', 'mobile': '9216360529', 'alter_mobile': '9216360529', 'landline': '9216360529'},
#  {'id': 5, 'RFPName': ('1B', '1C', '1D', 3), 'RFPMarks': (3, 1, 7, 20), 'TotalMarks': 31, 'name': 'Anjani Kumar',  'email': 'anjani.kumar@ictonline.com', 'state': 'BIHAR', 'district': 'Patna', 'address': 'D-80 Manju Niwas, Sai Netralaya, P. c.  Colony, kankarbagh, Patna, Bihar-80020', 'pan': 'AKNPK6116A', 'mobile': '7070761258', 'alter_mobile': '0', 'landline': '40863000'},
#  {'id': 1, 'RFPName': ('1B', '1C', 3), 'RFPMarks': (3, 1, 25), 'TotalMarks': 29, 'name': 'santosh kumar shrivastava',  'email': 'santosh@srivastavas.in', 'state': 'UTTAR PRADESH', 'district': 'Ghaziabad', 'address': 'STC-101,SUN TOWER C BLOCK,SHIPRA SUN CITY,INDIRAPURAM,GHAZIABAD-201014,', 'pan': 'ARIPS4696K', 'mobile': '8800297444', 'alter_mobile': '9871010187', 'landline': '01204104717'},
#  {'id': 4, 'RFPName': ('1B', '1C'), 'RFPMarks': (3, 1), 'TotalMarks': 4, 'name': 'Surya VIjay Pratap',  'email': 'pratapsv73@yahoo.com', 'state': 'HARYANA', 'district': 'Faridabad', 'address': 'Flat no 102 ,Exel Tower ,SPR Buildtech\r\nTigaon Road , Near BPTP, Sector 82\r\nFaridabad ,Pin 121007\r\n', 'pan': 'AJUPP1903G', 'mobile': '9810865160', 'alter_mobile': '0', 'landline': '01294140080'},
#  {'id': 6, 'RFPName': ('1B', '1C'), 'RFPMarks': (3, 1), 'TotalMarks': 4, 'name': ' Shailendra kumar',  'email': 'shailendra1969k@gmail.com', 'state': 'BIHAR', 'district': 'Patna', 'address': 'Alipur Bihta, Ward No.-10, Shiv Tample, Bihta-803202.', 'pan': '', 'mobile': '', 'alter_mobile': '', 'landline': ''}]
def ftnWriteRecordsToExcel(xlsFile, outputSheet, list_Consultant):
    try:
        if ""==xlsFile:
            xlsFile = 'C:\CVHUNT_SA\example.xlsx' #setting default XLS Locn
        if ""==outputSheet:
            outputSheet = "RFP Results"
        if(None == list_Consultant ):
            raise Exception('writingToExcel.py :::: list_Consultant in NULL')
        wb = load_workbook(xlsFile, False)

        print("writingToExcel.py-----------------  Sheet Data -------------- ")
        try:
            #Delete sheet named RFP Results
            #print("++++++++++++++",wb.sheetnames("RFP Results"))
            sheetnames=wb.sheetnames
            for sheet in sheetnames:
                if sheet == outputSheet:#Check if Sheet name exists in the List, if yes, get its reference and delete it
                    #print("GGGGGGGGGGGGGGGGGG",sheet)
                    refSheet=wb.get_sheet_by_name(outputSheet)
                    wb.remove_sheet(refSheet)
                    wb.save(xlsFile)

        except Exception as err:
            print("writeRecordsInExcel.py [[[[1]]]Exception occured:::unable_to_remove_sheet ", err)

        lastCellInExcel = 2 + list_Consultant.__len__()
        lastCellNo = 'M' + (str(lastCellInExcel))
        wsMdl = wb.create_sheet(outputSheet)
        # outputsheet.title ="CV Search Results for RFP"
        wsMdl['A1'] = "RFP Results"
        #wsMdl['A1'].style = 'Title'
        wsMdl.merge_cells('A1:N1')

        wsMdl['A1'].alignment = Alignment(horizontal='center', vertical='center')
        wsMdl['A1'].font = Font(b=True,underline='single', sz=20, color='0020C2')#Midnight color
        for rowno, rowOfCellObjects in enumerate(wsMdl['A2':'M2']): #Heading ROW
            for cellno, cellObj in enumerate(rowOfCellObjects):
                cellObj.alignment = Alignment(horizontal='center', vertical='center')
                cellObj.font= Font(bold=True, sz=14, color='1F45FC')#Blue Orchid

        wsMdl['A2'] = "Consultant ID"
        wsMdl['B2'] = "Consultant Name"
        wsMdl['C2'] = "RFP Names"
        wsMdl['D2'] = "RFP Marks"
        wsMdl['E2'] = "Total Marks"
        wsMdl['F2'] = "Address"
        wsMdl['G2'] = "District"
        wsMdl['H2'] = "State"
        wsMdl['I2'] = "EMail"
        wsMdl['J2'] = "Mobile No."
        wsMdl['K2'] = "Alternative Mobile"
        wsMdl['L2'] = "Landline"
        wsMdl['M2'] = "PAN Detail"


        print(lastCellInExcel," celllNO ",lastCellNo)
        #for row in mysheet.iter_rows(min_row=4, min_col=1, max_row=9, max_col=lastCellNo):
        for rowno, rowOfCellObjects in enumerate(wsMdl['A3':lastCellNo]):
            for cellno, cellObj in enumerate(rowOfCellObjects):
                print(rowno,"<<n record==>> ",list_Consultant[rowno],"cellObj.column=",cellObj.column,cellObj.row)
                print("Mdl sheet ** ", cellObj.coordinate, cellObj.value)
                if(cellObj.column == 'A'):
                    cellObj.alignment = Alignment(horizontal='center', vertical='center')
                    cellObj.value=list_Consultant[rowno]['id']
                elif(cellObj.column == 'B'):
                    cellObj.font = Font(color="008000", bold=True, sz=11)
                    cellObj.value=list_Consultant[rowno]['name']
                elif (cellObj.column == 'C'):
                    cellObj.value = str(list_Consultant[rowno]['RFPName']).strip('(').strip(")")
                elif (cellObj.column == 'D'):
                    cellObj.value = str(list_Consultant[rowno]['RFPMarks']).strip('(').strip(")")
                elif (cellObj.column == 'E'):
                    cellObj.alignment = Alignment(horizontal='center', vertical='center')
                    cellObj.font = Font(color='000080',bold=True,sz=11)#Navy Blue
                    cellObj.value = list_Consultant[rowno]['TotalMarks']
                elif (cellObj.column == 'F'):
                    cellObj.value = list_Consultant[rowno]['address']
                elif (cellObj.column == 'G'):
                    cellObj.value = list_Consultant[rowno]['district']
                elif (cellObj.column == 'H'):
                    cellObj.value = list_Consultant[rowno]['state']
                elif (cellObj.column == 'I'):
                    cellObj.value = list_Consultant[rowno]['email']
                elif(cellObj.column =='J'):
                    cellObj.value = list_Consultant[rowno]['mobile']
                elif(cellObj.column =='K'):
                    cellObj.value = list_Consultant[rowno]['alter_mobile']
                elif(cellObj.column =='L'):
                    cellObj.value = list_Consultant[rowno]['landline']
                elif (cellObj.column == 'M'):
                    cellObj.value = list_Consultant[rowno]['pan']
        dims = {}
        for row in wsMdl.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            wsMdl.column_dimensions[col].width = value + 5
        wb.save(xlsFile)


    except KeyError as err:
        print("Exception occured:::",err)