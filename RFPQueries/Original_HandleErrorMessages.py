from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, colors, Color
import traceback

#def PrintErrorMessages(msg,xls,sheetName,sheet,cell,type):
def PrintErrorMessages(errMsg,errCell,xls,sheetName,type):
    wb = load_workbook(xls, False)
    cell ='A20'
    cellCol='A'
    cellRow=20
    print("HandleErrorMessages.py-----------------  Sheet Data -------------- ",xls)
    try:
        print("++++++   HandleErrorMessages ++++++++")
        sheetnames = wb.sheetnames
        for name in sheetnames:
            print("|",name,"|","name","|",sheetName,"|")
            if name == sheetName:  # Check if Sheet name exists in the List, if yes, get its reference and delete it
                print("GGGGGGGGGGGGGGGGGG CHAR",sheetName)
                ws= wb.get_sheet_by_name(sheetName)
        if None == ws:
            print("HHHHHHHHHHHHHHHHHHHHH CHAR",sheetName)
            ws=wb.create_sheet(sheetName)
        print("444444444444444444444444   4444444444444",sheetName,ws['A1'].value)
        #sheet['A1'].value="msg[0] CHHAHSAHS"
        next=""
        for n, msg in  enumerate(errMsg):
            #_ = ws.cell(column=cellCol, row=cellRow, value=msg)
            ws[cell].alignment =ws[cell].alignment.copy(wrapText=True)
            #nextCell=ws._get_cell(row=cellRow+n,column=cellCol)
            #ws[cell+next()].value+="\n\n"+msg
            #nextCell.value=msg
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell].font = Font(bold=True, sz=14, color='FF0000', name='Arial')  # Maroon
            ws.cell(row=cellRow, column=cellCol).value=msg
            ws[errCell[n]].style =  'Warning Text'
            #ws.cell(row=cellRow, column=cellCol + n).alignment = Alignment(horizontal='center', vertical='center')
            #ws.cell(row=cellRow, column=cellCol + n).font= Font(bold=True, sz=14, color='FF0000', name='Arial')#Maroon
            print(n,"====5555555555555555",msg)
            import re
            ws.column_dimensions[cellCol].width = len(msg) + 2

        wb.save(xls)
        print("6666666666666666666666   6666666666666  6666666666666666666666")
    except Exception as err:
        traceback.print_tb(err.__traceback__)


        #print("HandleErrorMessages.py [[[[1]]]Exception occured::: ", err)
# print("charu")
# PrintErrorMessages("sbb","C:/Users/Paritosh.Paritosh-PC/PycharmProjects/CVHunt_SA/ExcelTemplate/example.xlsx","RFP1 - Qualification",'A20',"error")
# print("cxguota")