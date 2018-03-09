# Import pandas
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
# Load csv
#df = pd.read_csv("example.xls")
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('C:\AgilyticsCharu\PyExe-Charu\ExcelTemplate\example.xlsx',False)

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name
try:
    outputsheet = wb.get_sheet_by_name("RFP Results")
except KeyError as err:
    print("Exception occured:::",err)
    outputsheet = wb.create_sheet("RFP Results")
except Exception as err:
    print("Exception occured:::",err)


# Print the sheet title
print("outpus  outputsheet.title",outputsheet.title)

# Get currently active sheet
mysheet = wb.get_sheet_by_name("Sheet1")#wb.active

# Check `anotherSheet`
print("anotherSheet-->",mysheet.title)

# Retrieve the value of a certain cell
print("A1==>>",mysheet['A1'].value)
print("mysheet.cell(2,3).value",mysheet.cell(row=2, column=2).value)
mysheet.cell(row=2, column=2).value=34.580

# Select element 'B2' of your sheet
c = outputsheet['B2']

# Retrieve the row number of your element
print(c.row)

# Retrieve the column letter of your element
print(c.column)

# Retrieve the coordinates of the cell
print(c.coordinate)

# Convert Sheet to DataFrame
df = pd.DataFrame(mysheet.values)

#printing skipping/alternate columns in range 1 to 8
for i in range(1, 8, 2):
    print("***** datat===>>",i, mysheet.cell(row=i, column=1).value)

print("-----------------  Sheet Data -------------- ")
for rowOfCellObjects in mysheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')


print("dfdfdfdfdfdfdf df  dfdfdf-->",df.all)

# Writing Data to Output Sheet "RFP Results"

#outputsheet.title ="CV Search Results for RFP"
outputsheet['A1']="Employee Code"
outputsheet['B1']="Employee Name"
outputsheet['C1']="No Of Projects"
outputsheet['D1']="RFP Marks"

outputsheet['A2']=1
outputsheet['B2']="ch"
outputsheet['C2']=31
import datetime
outputsheet['D2']=datetime.datetime.now()
outputsheet.append=(["d","ch",1])


print(outputsheet.columns)
for col in outputsheet.columns:
    print("col==>", col, col)
    if col == 1:
        col="ssss"

print("done writing")
#df1 = pd.DataFrame
#df1.add_prefix()

ws = wb.get_sheet_by_name("Sheet2")
# Adding df values to a sheet "Sheet2"
# Append the rows of the DataFrame to your worksheet
#for r in dataframe_to_rows(df, index=True, header=True):
for r in dataframe_to_rows(df,index=False,header=False):
    ws.append(r)

print("-----------------  Sheet Data -------------- ")
wsMdl = wb.create_sheet("Model Sheet")
for rowno, rowOfCellObjects in enumerate(wsMdl['A1':'C3']):
    for cellno, cellObj in enumerate(rowOfCellObjects):
        print("Mdl sheet ** ",cellObj.coordinate, cellObj.value)
        cellObj.value = str(rowno)+", "+str(cellno)
    print('--- END OF ROW ---')
tu=(1,"charu",4.5,"hello")
wsMdl['D1']=str(tu)

print("Read data from a particular column")
#for row in mysheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
for row in mysheet.iter_rows(min_row=1, min_col=2, max_row=9, max_col=2):
    print("mysheet col 1 data ****")#,cellObj.value)
    for cell in row:
        print(cell.value, end=" ")

wb.save('C:\AgilyticsCharu\PyExe-Charu\ExcelTemplate\example.xlsx')