# Import pandas
import pandas as pd

# Load csv
#df = pd.read_csv("example.xls")

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./example.xlsx')

# Get sheet names
print(wb.get_sheet_names())

# Get a sheet by name
sheet = wb.get_sheet_by_name('Sheet3')

# Print the sheet title
sheet.title

# Get currently active sheet
anotherSheet = wb.active

# Check `anotherSheet`
print("anotherSheet-->",anotherSheet.title)

# Retrieve the value of a certain cell
print("A1==>>",anotherSheet['A1'].value)

# Select element 'B2' of your sheet
c = sheet['B2']

# Retrieve the row number of your element
print(c.row)

# Retrieve the column letter of your element
print(c.column)

# Retrieve the coordinates of the cell
print(c.coordinate)

# Convert Sheet to DataFrame
df = pd.DataFrame(anotherSheet.values)

print("dfdfdfdfdfdfdf df  dfdfdf-->",df.all)