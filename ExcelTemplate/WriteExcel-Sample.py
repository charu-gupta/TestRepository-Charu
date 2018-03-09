from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 142

# Rows can also be appended
ws.append([12, 212, "dhcharu3 "])
ws.append(["xxxxxxx", "yyyyyyy", "ZZZZZZZZZZZZZZZZZZZz", "AAAAAAAAAA","2313,ee,34",'er23,7'])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample_"+str(datetime.datetime.now().timestamp())+".xlsx")